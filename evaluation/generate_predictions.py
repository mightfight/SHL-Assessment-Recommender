"""
Generate predictions on the unlabeled test set.
Outputs Firstname_Lastname.csv in the required format:
  query, Assessment_url
"""

import os
import sys
import csv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import openpyxl
from recommender import recommend

DATASET_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Gen_AI Dataset.xlsx")
# The user needs firstname_lastname.csv, let's just make it output to project root
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Firstname_Lastname.csv")


def load_test_queries(filepath: str) -> list[str]:
    """Load unlabeled test queries from xlsx."""
    wb = openpyxl.load_workbook(filepath)

    # Try to find test sheet
    test_sheet = None
    for name in wb.sheetnames:
        if "test" in name.lower():
            test_sheet = wb[name]
            break
    if test_sheet is None and len(wb.sheetnames) > 1:
        test_sheet = wb[wb.sheetnames[1]]
    elif test_sheet is None:
        test_sheet = wb[wb.sheetnames[0]]

    print(f"Using sheet: '{test_sheet.title}'")

    rows = list(test_sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).lower().strip() if h else "" for h in rows[0]]
    query_col = next((i for i, h in enumerate(header) if "query" in h), 0)

    queries = []
    seen = set()
    for row in rows[1:]:
        if not row or not row[query_col]:
            continue
        q = str(row[query_col]).strip()
        if q and q not in seen:
            queries.append(q)
            seen.add(q)

    return queries


def generate_predictions():
    print("=" * 60)
    print("Generating Predictions on Test Set")
    print("=" * 60)

    queries = load_test_queries(DATASET_FILE)
    print(f"\nLoaded {len(queries)} test queries\n")

    rows = []
    for i, query in enumerate(queries):
        print(f"[{i+1}/{len(queries)}] {query[:80]}...")
        try:
            recs = recommend(query, num_results=10)
            for r in recs:
                rows.append({"query": query, "Assessment_url": r["url"]})
            print(f"  -> {len(recs)} recommendations")
        except Exception as e:
            print(f"  ERROR: {e}")

    # Write CSV
    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["query", "Assessment_url"])
        writer.writeheader()
        writer.writerows(rows)

    print(f"\n[SUCCESS] Predictions written to: {OUTPUT_FILE}")
    print(f"   Total rows: {len(rows)}")


if __name__ == "__main__":
    generate_predictions()
