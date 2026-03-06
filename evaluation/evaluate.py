"""
Evaluation: Mean Recall@K
Evaluates the recommendation pipeline against the labeled train set in Gen_AI Dataset.xlsx
"""

import os
import sys
import json

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import openpyxl
from recommender import recommend

DATASET_FILE = os.path.join(os.path.dirname(__file__), "..", "Gen_AI Dataset.xlsx")
K = 10  # Recall@K


def load_train_data(filepath: str) -> list[dict]:
    """Load labeled train set from xlsx. Returns list of {query, relevant_urls}."""
    wb = openpyxl.load_workbook(filepath)

    # Try sheet names: train, Train, labelled, Sheet1, etc.
    train_sheet = None
    for name in wb.sheetnames:
        if "train" in name.lower() or "label" in name.lower():
            train_sheet = wb[name]
            break
    if train_sheet is None:
        train_sheet = wb[wb.sheetnames[0]]

    print(f"Using sheet: '{train_sheet.title}' (rows={train_sheet.max_row})")

    rows = list(train_sheet.iter_rows(values_only=True))
    # First row is header
    if not rows:
        return []

    header = [str(h).lower().strip() if h else "" for h in rows[0]]
    query_col = next((i for i, h in enumerate(header) if "query" in h), 0)
    url_col = next((i for i, h in enumerate(header) if "url" in h or "assessment" in h), 1)

    # Group rows by query
    data = {}
    for row in rows[1:]:
        if not row or not row[query_col]:
            continue
        q = str(row[query_col]).strip()
        u = str(row[url_col]).strip() if row[url_col] else ""
        if q and u:
            data.setdefault(q, [])
            if u not in data[q]:
                data[q].append(u)

    return [{"query": q, "relevant_urls": urls} for q, urls in data.items()]


def recall_at_k(predicted_urls: list[str], relevant_urls: list[str], k: int = 10) -> float:
    """Compute Recall@K for a single query."""
    if not relevant_urls:
        return 0.0
    top_k = set(predicted_urls[:k])
    hits = sum(1 for url in relevant_urls if url in top_k)
    return hits / len(relevant_urls)


def normalize_url(url: str) -> str:
    """Normalize URL for comparison (handle /products/ vs /solutions/products/)."""
    url = url.strip().rstrip("/").lower()
    url = url.replace("https://www.shl.com/products/product-catalog/view/",
                       "https://www.shl.com/solutions/products/product-catalog/view/")
    return url


def evaluate():
    print("=" * 60)
    print("SHL Recommendation Evaluation — Mean Recall@10")
    print("=" * 60)

    train_data = load_train_data(DATASET_FILE)
    print(f"\nLoaded {len(train_data)} labeled queries\n")

    recalls = []
    results_log = []

    for i, item in enumerate(train_data):
        query = item["query"]
        relevant = [normalize_url(u) for u in item["relevant_urls"]]

        print(f"[{i+1}/{len(train_data)}] Query: {query[:80]}...")
        print(f"  Ground truth: {len(relevant)} relevant assessments")

        try:
            recs = recommend(query, num_results=K)
            predicted = [normalize_url(r["url"]) for r in recs]
        except Exception as e:
            print(f"  ERROR: {e}")
            predicted = []

        r_at_k = recall_at_k(predicted, relevant, k=K)
        recalls.append(r_at_k)

        hits = [u for u in relevant if u in set(predicted)]
        print(f"  Predicted: {len(predicted)} assessments")
        print(f"  Hits: {len(hits)}")
        print(f"  Recall@{K}: {r_at_k:.4f}")

        results_log.append({
            "query": query,
            "recall_at_k": r_at_k,
            "hits": len(hits),
            "relevant_count": len(relevant),
            "predicted_count": len(predicted),
        })
        print()

    mean_recall = sum(recalls) / len(recalls) if recalls else 0.0
    print("=" * 60)
    print(f"Mean Recall@{K}: {mean_recall:.4f}")
    print("=" * 60)

    # Save results log
    out_file = os.path.join(os.path.dirname(__file__), "eval_results.json")
    with open(out_file, "w") as f:
        json.dump({
            "mean_recall_at_k": mean_recall,
            "k": K,
            "per_query": results_log
        }, f, indent=2)
    print(f"\nDetailed results saved to: {out_file}")

    return mean_recall


if __name__ == "__main__":
    evaluate()
